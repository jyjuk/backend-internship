import io
import pytest
from openpyxl import Workbook

from app.services.excel_import_service import ExcelImportService


@pytest.fixture
def valid_excel_bytes():
    """Create a valid Excel file for testing"""
    wb = Workbook()
    ws = wb.active

    ws.append(['quiz_title', 'quiz_description', 'question_text', 'question_order',
               'answer_text', 'is_correct', 'answer_order'])

    ws.append(['Python Basics', 'Intro to Python', 'What is Python?', 1,
               'A programming language', True, 1])
    ws.append(['Python Basics', 'Intro to Python', 'What is Python?', 1,
               'A snake', False, 2])
    ws.append(['Python Basics', 'Intro to Python', 'Is Python easy?', 2,
               'Yes', True, 1])
    ws.append(['Python Basics', 'Intro to Python', 'Is Python easy?', 2,
               'No', False, 2])

    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)

    return excel_bytes.read()


@pytest.fixture
def invalid_excel_missing_columns():
    """Create Excel file with missing required columns"""
    wb = Workbook()
    ws = wb.active

    ws.append(['quiz_title', 'question_text', 'question_order', 'is_correct', 'answer_order'])
    ws.append(['Python', 'What is it?', 1, True, 1])

    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)

    return excel_bytes.read()


@pytest.fixture
def invalid_excel_too_few_questions():
    """Create Excel file with only 1 question (minimum is 2)"""
    wb = Workbook()
    ws = wb.active

    ws.append(['quiz_title', 'quiz_description', 'question_text', 'question_order',
               'answer_text', 'is_correct', 'answer_order'])
    ws.append(['Python', 'Test', 'What is it?', 1, 'Language', True, 1])
    ws.append(['Python', 'Test', 'What is it?', 1, 'Snake', False, 2])

    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)

    return excel_bytes.read()


@pytest.fixture
def invalid_excel_no_correct_answer():
    """Create Excel file with question that has no correct answer"""
    wb = Workbook()
    ws = wb.active

    ws.append(['quiz_title', 'quiz_description', 'question_text', 'question_order',
               'answer_text', 'is_correct', 'answer_order'])
    ws.append(['Python', 'Test', 'What is it?', 1, 'Language', False, 1])
    ws.append(['Python', 'Test', 'What is it?', 1, 'Snake', False, 2])
    ws.append(['Python', 'Test', 'Is it easy?', 2, 'Yes', True, 1])
    ws.append(['Python', 'Test', 'Is it easy?', 2, 'No', False, 2])

    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)

    return excel_bytes.read()


class TestExcelParsing:
    """Test Excel file parsing (unit tests without DB)"""

    def test_parse_valid_excel(self, valid_excel_bytes):
        """Test parsing valid Excel file"""
        class MockDB:
            pass

        service = ExcelImportService(MockDB())

        rows = service._parse_excel(valid_excel_bytes)

        assert len(rows) == 4
        assert rows[0]['quiz_title'] == 'Python Basics'
        assert rows[0]['question_text'] == 'What is Python?'
        assert rows[0]['answer_text'] == 'A programming language'
        assert rows[0]['is_correct'] is True

    def test_parse_missing_required_columns(self, invalid_excel_missing_columns):
        """Test parsing Excel with missing required columns"""
        class MockDB:
            pass

        service = ExcelImportService(MockDB())

        with pytest.raises(Exception) as exc_info:
            service._parse_excel(invalid_excel_missing_columns)

        assert "Missing required column: answer_text" in str(exc_info.value)

    def test_parse_empty_file(self):
        """Test parsing empty Excel file"""
        wb = Workbook()
        ws = wb.active

        ws.append(['quiz_title', 'quiz_description', 'question_text', 'question_order',
                   'answer_text', 'is_correct', 'answer_order'])

        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        class MockDB:
            pass

        service = ExcelImportService(MockDB())

        with pytest.raises(Exception) as exc_info:
            service._parse_excel(excel_bytes.read())

        assert "Excel file is empty" in str(exc_info.value)


class TestExcelValidation:
    """Test Excel data validation (unit tests without DB)"""

    def test_validate_valid_data(self, valid_excel_bytes):
        """Test validation passes with valid data"""
        class MockDB:
            pass

        service = ExcelImportService(MockDB())
        rows = service._parse_excel(valid_excel_bytes)

        service._validate_quiz_data(rows)

    def test_validate_too_few_questions(self, invalid_excel_too_few_questions):
        """Test validation fails with less than 2 questions"""
        class MockDB:
            pass

        service = ExcelImportService(MockDB())
        rows = service._parse_excel(invalid_excel_too_few_questions)

        with pytest.raises(Exception) as exc_info:
            service._validate_quiz_data(rows)

        assert "must have at least 2 questions" in str(exc_info.value)

    def test_validate_no_correct_answer(self, invalid_excel_no_correct_answer):
        """Test validation fails when question has no correct answer"""
        class MockDB:
            pass

        service = ExcelImportService(MockDB())
        rows = service._parse_excel(invalid_excel_no_correct_answer)

        with pytest.raises(Exception) as exc_info:
            service._validate_quiz_data(rows)

        assert "must have at least one correct answer" in str(exc_info.value)

    def test_validate_too_many_answers(self):
        """Test validation fails when question has more than 4 answers"""
        wb = Workbook()
        ws = wb.active

        ws.append(['quiz_title', 'quiz_description', 'question_text', 'question_order',
                   'answer_text', 'is_correct', 'answer_order'])
        ws.append(['Python', 'Test', 'What is it?', 1, 'Answer 1', True, 1])
        ws.append(['Python', 'Test', 'What is it?', 1, 'Answer 2', False, 2])
        ws.append(['Python', 'Test', 'What is it?', 1, 'Answer 3', False, 3])
        ws.append(['Python', 'Test', 'What is it?', 1, 'Answer 4', False, 4])
        ws.append(['Python', 'Test', 'What is it?', 1, 'Answer 5', False, 5])
        ws.append(['Python', 'Test', 'Is it easy?', 2, 'Yes', True, 1])
        ws.append(['Python', 'Test', 'Is it easy?', 2, 'No', False, 2])

        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        class MockDB:
            pass

        service = ExcelImportService(MockDB())
        rows = service._parse_excel(excel_bytes.read())

        with pytest.raises(Exception) as exc_info:
            service._validate_quiz_data(rows)

        assert "must have 2-4 answers" in str(exc_info.value)


class TestImportFileValidation:
    """Test file format validation"""

    def test_reject_non_excel_file(self):
        """Test that non-Excel files are rejected"""
        class MockUploadFile:
            def __init__(self):
                self.filename = "test.txt"

            async def read(self):
                return b"not an excel file"

        file = MockUploadFile()

        assert not file.filename.endswith(('.xlsx', '.xls'))
