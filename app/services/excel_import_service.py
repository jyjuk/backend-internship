"""
Excel Import Service for Quiz Management
"""
import io
from typing import List, Dict, Any, Optional
from uuid import UUID
from fastapi import UploadFile, HTTPException
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.quiz import Quiz
from app.models.question import Question
from app.models.answer import Answer
from app.repositories.quiz import QuizRepository
from app.repositories.question import QuestionRepository
from app.repositories.answer import AnswerRepository
from app.services.quiz_service import QuizService


class ExcelImportService:
    """Service for importing quizzes from Excel files"""

    def __init__(self, db: AsyncSession):
        self.db = db
        self.quiz_repo = QuizRepository(db)
        self.question_repo = QuestionRepository(db)
        self.answer_repo = AnswerRepository(db)
        self.quiz_service = QuizService(db)

    async def import_quizzes_from_excel(
            self,
            file: UploadFile,
            company_id: UUID,
            user_id: UUID
    ) -> Dict[str, Any]:
        """
        Import quizzes from Excel file

        Args:
            file: Uploaded Excel file
            company_id: Company ID
            user_id: Current user ID (for permission check)

        Returns:
            Dict with import results (created, updated, errors)
        """
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=400,
                detail="Invalid file format. Please upload Excel file (.xlsx or .xls)"
            )

        await self.quiz_service._check_owner_or_admin(company_id, user_id)

        contents = await file.read()

        quiz_data = self._parse_excel(contents)

        self._validate_quiz_data(quiz_data)

        result = await self._process_quizzes(quiz_data, company_id)

        return result

    def _parse_excel(self, contents: bytes) -> List[Dict[str, Any]]:
        """
        Parse Excel file and extract quiz data

        Expected columns:
        - quiz_title (required)
        - quiz_description (optional)
        - question_text (required)
        - question_order (required)
        - answer_text (required)
        - is_correct (required: TRUE/FALSE)
        - answer_order (required)
        """
        try:
            workbook = load_workbook(io.BytesIO(contents))
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]

            required_columns = [
                "quiz_title", "question_text", "question_order",
                "answer_text", "is_correct", "answer_order"
            ]

            for col in required_columns:
                if col not in headers:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Missing required column: {col}"
                    )

            rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                row_dict = dict(zip(headers, row))
                rows.append(row_dict)

            if not rows:
                raise HTTPException(
                    status_code=400,
                    detail="Excel file is empty. Please add quiz data."
                )

            return rows

        except Exception as e:
            if isinstance(e, HTTPException):
                raise
            raise HTTPException(
                status_code=400,
                detail=f"Error parsing Excel file: {str(e)}"
            )

    def _validate_quiz_data(self, rows: List[Dict[str, Any]]) -> None:
        """Validate parsed quiz data"""

        quizzes = {}
        for row in rows:
            quiz_title = row.get('quiz_title', '').strip()

            if not quiz_title:
                raise HTTPException(
                    status_code=400,
                    detail="quiz_title cannot be empty"
                )

            if quiz_title not in quizzes:
                quizzes[quiz_title] = {
                    'questions': {}
                }

            question_text = row.get('question_text', '').strip()
            if not question_text:
                raise HTTPException(
                    status_code=400,
                    detail=f"question_text cannot be empty for quiz '{quiz_title}'"
                )

            question_order = row.get('question_order')
            if question_order is None:
                raise HTTPException(
                    status_code=400,
                    detail=f"question_order is required for question '{question_text}'"
                )

            answer_text = row.get('answer_text', '').strip()
            if not answer_text:
                raise HTTPException(
                    status_code=400,
                    detail=f"answer_text cannot be empty for question '{question_text}'"
                )

            is_correct = row.get("is_correct")
            if is_correct is None:
                raise HTTPException(
                    status_code=400,
                    detail=f"is_correct is required for answer '{answer_text}'"
                )

            answer_order = row.get("answer_order")
            if answer_order is None:
                raise HTTPException(
                    status_code=400,
                    detail=f"answer_order is required for answer '{answer_text}'"
                )

            question_key = f"{question_text}_{question_order}"
            if question_key not in quizzes[quiz_title]["questions"]:
                quizzes[quiz_title]["questions"][question_key] = {
                    "text": question_text,
                    "order": question_order,
                    "answers": []
                }

            quizzes[quiz_title]["questions"][question_key]["answers"].append({
                "text": answer_text,
                "is_correct": bool(is_correct),
                "order": answer_order
            })

        for quiz_title, quiz_data in quizzes.items():
            questions = quiz_data["questions"]

            if len(questions) < 2:
                raise HTTPException(
                    status_code=400,
                    detail=f"Quiz '{quiz_title}' must have at least 2 questions"
                )

            for question_key, question in questions.items():
                answer_count = len(question["answers"])
                if answer_count < 2 or answer_count > 4:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Question '{question['text']}' must have 2-4 answers (has {answer_count})"
                    )

                correct_count = sum(1 for a in question["answers"] if a["is_correct"])
                if correct_count == 0:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Question '{question['text']}' must have at least one correct answer"
                    )

    async def _process_quizzes(
            self,
            rows: List[Dict[str, Any]],
            company_id: UUID
    ) -> Dict[str, Any]:
        """Process and import quizzes"""

        quizzes_data = {}
        for row in rows:
            quiz_title = row["quiz_title"].strip()

            if quiz_title not in quizzes_data:
                quizzes_data[quiz_title] = {
                    "title": quiz_title,
                    "description": row.get("quiz_description", "").strip() if row.get("quiz_description") else None,
                    "questions": {}
                }

            question_text = row["question_text"].strip()
            question_order = int(row["question_order"])
            question_key = f"{question_text}_{question_order}"

            if question_key not in quizzes_data[quiz_title]["questions"]:
                quizzes_data[quiz_title]["questions"][question_key] = {
                    "title": question_text,
                    "order": question_order,
                    "answers": []
                }

            quizzes_data[quiz_title]["questions"][question_key]["answers"].append({
                "text": row["answer_text"].strip(),
                "is_correct": bool(row["is_correct"]),
                "order": int(row["answer_order"])
            })

        created_count = 0
        updated_count = 0
        errors = []

        for quiz_title, quiz_data in quizzes_data.items():
            try:
                existing_quiz = await self.quiz_repo.get_by_title_and_company(
                    title=quiz_data["title"],
                    company_id=company_id
                )

                if existing_quiz:
                    await self._update_quiz(existing_quiz, quiz_data)
                    updated_count += 1
                else:
                    await self._create_quiz(quiz_data, company_id)
                    created_count += 1

            except Exception as e:
                errors.append(f"Error processing quiz '{quiz_title}': {str(e)}")

        await self.db.commit()

        return {
            "created": created_count,
            "updated": updated_count,
            "total": created_count + updated_count,
            "errors": errors
        }

    async def _create_quiz(self, quiz_data: Dict[str, Any], company_id: UUID) -> Quiz:
        """Create new quiz with questions and answers"""

        quiz = Quiz(
            company_id=company_id,
            title=quiz_data["title"],
            description=quiz_data["description"]
        )
        self.db.add(quiz)
        await self.db.flush()

        for question_data in quiz_data["questions"].values():
            question = Question(
                quiz_id=quiz.id,
                title=question_data["title"],
                order=question_data["order"]
            )
            self.db.add(question)
            await self.db.flush()

            for answer_data in question_data["answers"]:
                answer = Answer(
                    question_id=question.id,
                    text=answer_data["text"],
                    is_correct=answer_data["is_correct"],
                    order=answer_data["order"]
                )
                self.db.add(answer)

        return quiz

    async def _update_quiz(self, existing_quiz: Quiz, quiz_data: Dict[str, Any]) -> Quiz:
        """Update existing quiz (replace questions and answers)"""

        existing_quiz.title = quiz_data["title"]
        if quiz_data["description"]:
            existing_quiz.description = quiz_data["description"]

        await self.question_repo.delete_by_quiz_id(existing_quiz.id)
        await self.db.flush()

        for question_data in quiz_data["questions"].values():
            question = Question(
                quiz_id=existing_quiz.id,
                title=question_data["title"],
                order=question_data["order"]
            )
            self.db.add(question)
            await self.db.flush()

            for answer_data in question_data["answers"]:
                answer = Answer(
                    question_id=question.id,
                    text=answer_data["text"],
                    is_correct=answer_data["is_correct"],
                    order=answer_data["order"]
                )
                self.db.add(answer)

        return existing_quiz
